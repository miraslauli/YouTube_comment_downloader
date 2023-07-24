import json

import openpyxl
from googleapiclient.discovery import build


API_KEY = ""
youtube = build("youtube", "v3", developerKey=API_KEY)
channel_id = ""


def get_channel_comments(channel_id: str) -> list:
    comments = []
    nextPageToken = None

    while True:
        try:
            response = youtube.commentThreads().list(
                part=['id', 'snippet'],
                allThreadsRelatedToChannelId=channel_id,
                maxResults=100,
                pageToken=nextPageToken,
            ).execute()

            for item in response["items"]:
                video_id = item["snippet"]["videoId"]
                comment_info = {
                    "comment_id": item["id"],
                    "video_id": video_id,
                    "author_display_name": item["snippet"]["topLevelComment"]["snippet"]["authorDisplayName"],
                    "author_channel_url": item["snippet"]["topLevelComment"]["snippet"]["authorChannelUrl"],
                    "text_display": item["snippet"]["topLevelComment"]["snippet"]["textDisplay"],
                    "text_original": item["snippet"]["topLevelComment"]["snippet"]["textOriginal"],
                    "like_count": item["snippet"]["topLevelComment"]["snippet"]["likeCount"],
                    "published_at": item["snippet"]["topLevelComment"]["snippet"]["publishedAt"],
                    "viewer_rating": item["snippet"]["topLevelComment"]["snippet"]["viewerRating"],
                    "total_reply_count": item["snippet"]["totalReplyCount"],
                }
                comments.append(comment_info)

            nextPageToken = response.get("nextPageToken")
            if not nextPageToken:
                break

        except TimeoutError:
            print('Timeout, continue')
            continue

        except Exception as exc:
            print(f'Other excretion: {exc}')
            return comments

    return comments


def save_to_json(comments: list) -> None:
    with open("file.json", 'w', encoding='utf-8') as json_file:
        json.dump(comments, json_file, ensure_ascii=False, indent=4)


def save_comments_to_excel(comments: list, file_name: str) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    column_names = [
        "Comment ID",
        "Video ID",
        "Author Display Name",
        "Author Channel URL",
        "Text Display",
        "Text Original",
        "Like Count",
        "Published At",
        "Viewer Rating",
        "Total Reply Count",
    ]

    try:
        for col_idx, col_name in enumerate(column_names, start=1):
            sheet.cell(row=1, column=col_idx, value=col_name)

        for idx, comment in enumerate(comments, start=2):
            sheet.cell(row=idx, column=1, value=comment["comment_id"])
            sheet.cell(row=idx, column=2, value=comment["video_id"])
            sheet.cell(row=idx, column=3, value=comment["author_display_name"])
            sheet.cell(row=idx, column=4, value=comment["author_channel_url"])
            sheet.cell(row=idx, column=5, value=comment["text_display"])
            sheet.cell(row=idx, column=6, value=comment["text_original"])
            sheet.cell(row=idx, column=7, value=comment["like_count"])
            sheet.cell(row=idx, column=8, value=comment["published_at"])
            sheet.cell(row=idx, column=9, value=comment["viewer_rating"])
            sheet.cell(row=idx, column=10, value=comment["total_reply_count"])
        workbook.save(file_name)

    except Exception as exc:
        print(f'Something wrong: {exc}')
        workbook.save(file_name)


if __name__ == "__main__":
    comments = get_channel_comments(channel_id)
    save_to_json(comments)
    save_comments_to_excel(comments, "file.xlsx")
